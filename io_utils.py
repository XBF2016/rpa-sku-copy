from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from common import append_to_log
 
import time
import hashlib
from urllib.parse import urlparse
try:
    import requests  # 可选依赖，用于下载图片到本地
except Exception:
    requests = None  # type: ignore
try:
    from PIL import Image as PILImage
    _HAS_PILLOW = True
except Exception:
    PILImage = None  # type: ignore
    _HAS_PILLOW = False
try:
    import win32com.client as win32  # 可选依赖，用于通过 Excel COM 插入“链接的图片”
except Exception:
    win32 = None  # type: ignore

# 导出相关工具

def _px_to_col_width(px: float) -> float:
    """将像素近似换算为 Excel 列宽(字符数)（保留，供可能的复用）。"""
    try:
        return max(1.0, (float(px) - 5.0) / 7.0)
    except Exception:
        return 12.0

def _px_to_points(px: float) -> float:
    """像素 -> point（1 pt = 1/72 in, 96 DPI 假定）。"""
    return float(px) * 72.0 / 96.0


def _sanitize_filename_component(name: str) -> str:
    """清理文件名非法字符并压缩空白。"""
    try:
        s = str(name or "").strip()
    except Exception:
        s = ""
    # Windows 非法字符: \ / : * ? " < > |
    for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        s = s.replace(ch, ' ')
    # 压缩空白
    s = ' '.join(s.split())
    # 避免结尾空格或点
    s = s.rstrip(' .')
    # 简单截断，保留足够长度（考虑后缀）
    if len(s) > 120:
        s = s[:120]
    return s or "未命名"


def _generate_spec_image_filenames(headers: List[str], results: List[List[str]], out_dir: Path):
    """基于结果表推断每个规格图的友好文件名。
    返回 (spec_img_urls, url_to_filename)，其中：
      - spec_img_urls: 按首次出现顺序的唯一图片URL列表
      - url_to_filename: URL -> 友好文件名（如：[颜色分类]胡桃木床 柔光夜灯带公牛插座.png）
    规则：
      - 找到引用该 URL 的所有行，识别值恒定的维度列；
      - 若存在多个恒定维度，优先名字含“颜色/色/color”的维度；否则取第一个恒定维度；
      - 文件名 = f"[{维度名}]{选项}.png"；若无法识别，则回退为 f"图片_{md5前10}.png"；
      - 若同名冲突，则自动追加 " (2)", "(3)" 等后缀；
    """
    # 定位图片链接列索引
    img_link_idx = -1
    try:
        img_link_idx = headers.index("图片链接")
    except Exception:
        # 兼容：当 headers 只有“各维度 + 价格”，数据行中可能额外包含一个隐藏的图片链接列
        try:
            if headers and headers[-1] == "价格":
                # 常见形态：行长度 = len(headers) + 1，且该额外列为维度之后、价格之前
                expected_img_idx = max(0, len(headers) - 1)
                has_extra = any((len(r) == len(headers) + 1) for r in (results or []))
                if has_extra:
                    img_link_idx = expected_img_idx
                else:
                    # 兜底：扫描找出“看起来像URL”的稳定列
                    candidate_counts = {}
                    for row in results or []:
                        for j, v in enumerate(row):
                            try:
                                s = str(v).strip().lower()
                            except Exception:
                                s = ''
                            if s.startswith("http://") or s.startswith("https://"):
                                candidate_counts[j] = candidate_counts.get(j, 0) + 1
                    if candidate_counts:
                        img_link_idx = max(candidate_counts.items(), key=lambda kv: kv[1])[0]
        except Exception:
            img_link_idx = -1

    # 维度列头（若能定位图片链接列，则取其之前为维度；否则若以“价格”结尾，则维度=去掉最后一列）
    try:
        if img_link_idx != -1:
            dim_names = [str(h) for h in headers[:img_link_idx]] if img_link_idx > 0 else []
        else:
            dim_names = [str(h) for h in headers[:-1]] if (headers and headers[-1] == "价格") else []
    except Exception:
        dim_names = []
    dim_count = len(dim_names)

    # URL 去重（首次出现顺序）及被哪些行引用
    spec_img_urls: List[str] = []
    url_to_rows: dict[str, List[List[str]]] = {}
    for row in results or []:
        try:
            img_link = ''
            if img_link_idx != -1 and len(row) > img_link_idx:
                img_link = str(row[img_link_idx] or '').strip()
            if not img_link or not (img_link.startswith('http://') or img_link.startswith('https://')):
                continue
            if img_link not in url_to_rows:
                url_to_rows[img_link] = []
                spec_img_urls.append(img_link)
            url_to_rows[img_link].append(row)
        except Exception:
            continue

    def _prefer_dim_index(name: str) -> int:
        try:
            nm = (name or '').lower()
        except Exception:
            nm = ''
        if any(k in nm for k in ['颜色', '顏色', '色', 'color']):
            return 0
        return 1

    # 生成唯一文件名
    used_names: set[str] = set()
    url_to_filename: dict[str, str] = {}
    for url in spec_img_urls:
        rows_for_img = url_to_rows.get(url, [])
        chosen_dim_idx = None
        if rows_for_img:
            constant_dims: List[int] = []
            for di in range(dim_count):
                try:
                    values = {str(r[di]).strip() for r in rows_for_img}
                except Exception:
                    values = set()
                if len(values) == 1:
                    constant_dims.append(di)
            if constant_dims:
                chosen_dim_idx = sorted(constant_dims, key=lambda i: (_prefer_dim_index(dim_names[i]), i))[0]

        if chosen_dim_idx is not None:
            try:
                dim_name = _sanitize_filename_component(dim_names[chosen_dim_idx])
            except Exception:
                dim_name = '规格'
            try:
                opt_text = _sanitize_filename_component(rows_for_img[0][chosen_dim_idx])
            except Exception:
                opt_text = '未命名'
            base = f"[{dim_name}]{opt_text}"
        else:
            # 回退：使用短哈希
            base = f"图片_{hashlib.md5(url.encode('utf-8')).hexdigest()[:10]}"

        candidate = base
        suffix = 2
        # 仅在本次生成范围内去重（避免与历史同名文件冲突导致 YAML 出现 "(2)"）
        while True:
            filename = candidate + ".png"
            if filename not in used_names:
                used_names.add(filename)
                break
            candidate = f"{base} ({suffix})"
            suffix += 1
        url_to_filename[url] = filename

    return spec_img_urls, url_to_filename


def _insert_linked_images_via_excel(xlsx_path: Path, img_col_idx: int, url_to_filename: dict, results: List[List[str]]) -> None:
    """使用 Excel COM（win32com）在目标列插入“链接的图片”，不将图片数据保存进 Excel。
    - LinkToFile=True, SaveWithDocument=False
    - 形状随单元格移动/缩放，按单元格尺寸等比缩放并居中。
    """
    if img_col_idx == -1 or not results:
        return
    if win32 is None:
        try:
            print("[提示] 未安装 win32com（pywin32），将不会在表格中插入可见图片，仅保留链接列。")
        except Exception:
            pass
        try:
            append_to_log("未安装/不可用的 Excel COM，跳过插入‘链接的图片’，仅保留超链接")
        except Exception:
            pass
        return

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path))
        ws = wb.Worksheets(1)

        # 删除目标列（从第2行开始）已有的形状，避免重复叠加
        try:
            for i in range(ws.Shapes.Count, 0, -1):
                shp = ws.Shapes.Item(i)
                try:
                    if shp.TopLeftCell.Column == (img_col_idx + 1) and shp.TopLeftCell.Row >= 2:
                        shp.Delete()
                except Exception:
                    pass
        except Exception:
            pass

        # 逐行插入“链接的图片”
        inserted_total = 0
        inserted_via_file = 0
        inserted_via_url = 0
        for r_idx, row in enumerate(results, start=2):  # Excel 行号从2开始（第1行为表头）
            try:
                url = str(row[img_col_idx]).strip()
            except Exception:
                url = ""
            fname = url_to_filename.get(url)
            png_path = (xlsx_path.parent / fname).resolve() if fname else None

            cell = ws.Cells(r_idx, img_col_idx + 1)
            # 若该单元格处于合并区域中，且不是合并区域的首行，则跳过，避免重复插入同一张图片
            try:
                if cell.MergeCells:
                    try:
                        top_row = int(cell.MergeArea.Row)
                    except Exception:
                        top_row = int(cell.Row)
                    if int(cell.Row) != top_row:
                        continue
            except Exception:
                pass
            left = float(cell.Left)
            top = float(cell.Top)
            # 若为合并单元格，优先使用整个合并区域的宽高，确保图片能铺满合并后的单元格
            try:
                if cell.MergeCells:
                    area = cell.MergeArea
                    cell_w = float(area.Width)
                    cell_h = float(area.Height)
                else:
                    cell_w = float(cell.Width)
                    cell_h = float(cell.Height)
            except Exception:
                cell_w = float(cell.Width)
                cell_h = float(cell.Height)
            # 计算按原始比例缩放后适配单元格的尺寸（优先使用本地PNG的真实尺寸）
            target_w, target_h = cell_w, cell_h
            if png_path is not None and png_path.exists():
                try:
                    if _HAS_PILLOW:
                        with PILImage.open(str(png_path)) as _im:
                            ow, oh = _im.size
                        if ow > 0 and oh > 0 and cell_w > 0 and cell_h > 0:
                            sc = min(cell_w / float(ow), cell_h / float(oh), 1.0)
                            target_w = max(1.0, float(ow) * sc)
                            target_h = max(1.0, float(oh) * sc)
                except Exception:
                    target_w, target_h = cell_w, cell_h

            try:
                # 优先使用本地 PNG 文件（若存在且路径有效）
                if png_path is not None and png_path.exists():
                    shp = ws.Shapes.AddPicture(str(png_path), True, False, left, top, target_w, target_h)
                    inserted_via_file += 1
                else:
                    # 本地文件不可用：尝试直接使用网络 URL 插入（同样为“链接”方式）
                    if url and (url.startswith("http://") or url.startswith("https://")):
                        shp = ws.Shapes.AddPicture(str(url), True, False, left, top, cell_w, cell_h)
                        inserted_via_url += 1
                    else:
                        continue
            except Exception:
                # 兼容部分版本：尝试 AddPicture2
                try:
                    if png_path is not None and png_path.exists():
                        shp = ws.Shapes.AddPicture2(str(png_path), True, False, left, top, target_w, target_h)
                        inserted_via_file += 1
                    else:
                        if url and (url.startswith("http://") or url.startswith("https://")):
                            shp = ws.Shapes.AddPicture2(str(url), True, False, left, top, cell_w, cell_h)
                            inserted_via_url += 1
                        else:
                            continue
                except Exception:
                    continue

            try:
                # 随单元格移动与缩放
                shp.Placement = 1  # xlMoveAndSize
                shp.LockAspectRatio = True
                # 仅设置可读的替代文本，避免修改链接源导致某些环境下路径解析问题
                try:
                    shp.AlternativeText = png_path.name
                except Exception:
                    pass
                # 再次确保完全贴合单元格边界（等比缩放）
                if cell_w > 0 and cell_h > 0 and shp.Width and shp.Height:
                    scale = min(cell_w / max(1.0, float(shp.Width)), cell_h / max(1.0, float(shp.Height)), 1.0)
                    shp.Width = max(1.0, float(shp.Width) * scale)
                    shp.Height = max(1.0, float(shp.Height) * scale)
                # 居中对齐（至少有一边会完全贴合单元格边界，比例不变）。若为合并区域，则基于合并区域宽高进行居中。
                shp.Left = float(cell.Left) + max(0.0, (float(cell_w) - float(shp.Width)) / 2.0)
                shp.Top = float(cell.Top) + max(0.0, (float(cell_h) - float(shp.Height)) / 2.0)
            except Exception:
                pass
            else:
                inserted_total += 1

        wb.Save()
        wb.Close(SaveChanges=True)
        try:
            if inserted_via_file or inserted_via_url:
                print(f"[步骤] 已通过Excel COM插入图片: 本地文件 {inserted_via_file} 张，网络URL {inserted_via_url} 张")
        except Exception:
            pass
        try:
            if inserted_total:
                append_to_log(f"通过 Excel COM 插入图片完成：本地 {inserted_via_file}，网络 {inserted_via_url}，总计 {inserted_total}")
            else:
                append_to_log("Excel COM 插入图片数量为 0：可能需要在 Excel 中点击‘启用内容’或检查图片本地文件/网络可达性")
        except Exception:
            pass
    finally:
        try:
            excel.Quit()
        except Exception:
            pass

def _merge_consecutive_cells(ws, results: List[List[str]], headers: List[str], img_col_idx: int) -> None:
    """按列合并相邻且值相同的单元格（第1行为表头，从第2行开始）。
    - 对“图片”列：基于原始图片 URL 判断是否相同（而不是单元格显示的“查看图片”文本）。
    - 其他列：基于写入的值进行等值判断（去掉首尾空白；数值按原值）。
    """
    try:
        if not results or not headers:
            return
        n_rows = len(results)
        n_cols = len(headers)

        # 价格列不进行合并（通常为最后一列）
        price_col_idx = -1
        try:
            if headers and headers[-1] == "价格":
                price_col_idx = len(headers) - 1
        except Exception:
            price_col_idx = -1

        def get_key(row_idx0: int, col_idx0: int):
            try:
                val = results[row_idx0][col_idx0]
            except Exception:
                return ""
            if col_idx0 == img_col_idx:
                try:
                    return str(val).strip()
                except Exception:
                    return ""
            # 非图片列：数值原值，文本去空白
            if isinstance(val, (int, float)):
                return val
            try:
                return str(val).strip()
            except Exception:
                return ""

        for col in range(n_cols):
            if col == price_col_idx:
                continue
            start = 0  # results 的 0 基索引（对应 Excel 的第 2 行）
            prev_key = get_key(0, col) if n_rows > 0 else None
            for i in range(1, n_rows + 1):  # 走到 n_rows 作为哨兵触发收尾
                key = get_key(i, col) if i < n_rows else None
                same = (key == prev_key) and (key not in (None, ""))
                if same:
                    continue
                # 结束上一段
                seg_len = i - start
                if seg_len >= 2 and (prev_key not in (None, "")):
                    r1 = start + 2  # +1 表头 +1 从 0 到 Excel 行
                    r2 = i + 1
                    c = col + 1
                    try:
                        ws.merge_cells(start_row=r1, end_row=r2, start_column=c, end_column=c)
                    except Exception:
                        pass
                start = i
                prev_key = key
    except Exception:
        # 合并失败不影响导出
        pass

def export_results_to_excel(results: List[List[str]], headers: List[str], file_path: Path) -> Path:
    """使用 openpyxl 将结果写入 Excel（含表头），返回最终保存的文件路径（可能是带时间戳的备选文件）。
    行为说明：
    - Excel 仅导出“各维度 + 价格”。若结果行中包含隐藏的图片链接列，将自动忽略，不写入工作簿。
    - 全表统一样式：水平/垂直居中，自动换行；文本列近似自适应列宽。
    - 若最后一列表头为“价格”，将该列写为纯数值（去掉人民币符号与千分位）。
    - 自动合并：同一列中相邻且值相同的单元格会自动合并。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "结果"

    # 写入表头
    ws.append(headers)
    # 表头样式：居中 + 自动换行
    for c in ws[1]:
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 统计每列最大文本长度（用于后续列宽自适应）
    max_text_len = [len(str(h)) for h in headers]

    # 识别“图片”与“图片链接”列索引（没有则返回 -1）——当前逻辑已不包含图片列
    img_col_idx = -1
    img_link_col_idx = -1

    # 去重图片链接并下载到与 Excel 同级目录（统一转为 PNG，使用友好文件名如：[维度]选项.png）
    img_output_dir = file_path.parent
    url_to_filename = {}
    # 已移除图片处理逻辑

    # 统一设置图片与链接列列宽（若存在）
    # 不含图片列，无需设置图片列宽

    for row in results:
        # 价格列数值化（要求“价格”为最后一列表头）
        if headers and headers[-1] == "价格":
            # 价格在行的最后一个元素（隐藏图片列在价格前一位，故不能用表头索引）
            try:
                last_val = row[-1]
            except Exception:
                last_val = ""
            if any(ch.isdigit() for ch in str(last_val)):
                row[-1] = float(next((t for t in ''.join((c if (c.isdigit() or c in '.,') else ' ') for c in str(last_val)).split() if t), '0').replace(',', ''))

        # 写入“各维度 + 价格”：取前 len(headers)-1 个为维度，最后一项强制取行末尾（价格）
        try:
            dims_part = list(row[:max(0, len(headers) - 1)])
        except Exception:
            dims_part = []
        try:
            price_part = [row[-1]] if headers and len(headers) >= 1 else []
        except Exception:
            price_part = [""]
        row_for_write = dims_part + price_part
        ws.append(row_for_write)
        current_row = ws.max_row

        # 当前行单元格统一样式（上下左右居中、自动换行）
        for ci in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=ci).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 统计文本长度用于自适应列宽（以写入后的内容计算，图片列跳过）
        for j in range(len(headers)):
            try:
                txt = row_for_write[j] if j < len(row_for_write) else ""
                txt_len = len(str(txt))
            except Exception:
                txt_len = 0
            if j < len(max_text_len) and txt_len > max_text_len[j]:
                max_text_len[j] = txt_len

        # 不含图片列，无需写入备用超链接

    # 文本列近似自适应列宽（跳过“图片”与“图片链接”列）
    for j in range(len(headers)):
        col_letter_j = get_column_letter(j + 1)
        # 控制在 [10, 40] 以内，避免过窄或过宽
        target_width = max(10, min(40, (max_text_len[j] if j < len(max_text_len) else 10) + 2))
        ws.column_dimensions[col_letter_j].width = target_width

    # 若存在图片列，给数据行一个适中的行高，便于后续 COM 插入的图片按单元格自适应可见
    # 不含图片列，无需设置行高

    # 保存前：按列合并相邻相同值（包含“图片”列基于URL判断）
    try:
        _merge_consecutive_cells(ws, results, headers, -1)
    except Exception:
        pass

    # 保存工作簿；若被占用则自动降级另存为 result_YYYYMMDD_HHMMSS.xlsx
    final_path = file_path
    try:
        wb.save(str(final_path))
    except Exception as e:
        try:
            ts = time.strftime("%Y%m%d_%H%M%S")
            alt = final_path.with_name(f"{final_path.stem}_{ts}{final_path.suffix}")
            wb.save(str(alt))
            final_path = alt
            print(f"[警告] Excel 文件被占用或写入失败，已改为另存为: {alt}")
        except Exception as e2:
            print(f"[错误] 导出Excel失败: {e2}")
            raise
    
    # 不再插入图片
    return final_path


def export_results_to_yaml(sku_dimensions, results: List[List[str]], headers: List[str], file_path: Path) -> Path:
    """将规格维度与组合信息导出为更清晰的 YAML：
    - specs: 维度 -> 选项列表（与示例一致，便于人工查看）
    - dims:  维度名称顺序（用于索引解释）
    - product_images: 商品主图画廊（当前未专门采集，暂为空数组，后续可扩展）
    - product_images_local: 与 product_images 对齐的本地文件名（当前为空数组）
    - spec_images: 规格图列表（主图区域在选中带图规格后展示的图片），每项为：
        - file: 本地文件名（如：[颜色分类]米白色.png）
        - url: 图片URL
    - combos: 紧凑列表，每行 = [每个维度的选项索引..., 价格]
      说明：
        - 选项索引为基于 specs[维度] 中的序号（从 0 开始）
        - 价格为数值（若无法解析则为 null）
    """
    # 计算维度名称列表（与 headers 对齐）
    try:
        if headers and headers[-1] == "价格":
            dim_names = [str(h) for h in headers[:-1]]
        else:
            dim_names = [getattr(d, 'name', f'维度{i+1}') for i, d in enumerate(sku_dimensions)]
    except Exception:
        dim_names = [getattr(d, 'name', f'维度{i+1}') for i, d in enumerate(sku_dimensions)]
    dim_count = len(dim_names)

    # 构建 specs 映射：维度 -> 选项列表，并保留维度顺序
    specs_map = {}
    try:
        for d in sku_dimensions:
            name = str(getattr(d, 'name', '') or '')
            opts = []
            try:
                for o in getattr(d, 'options', []) or []:
                    txt = str(getattr(o, 'text', '') or '').strip()
                    if txt:
                        opts.append(txt)
            except Exception:
                pass
            if name and opts:
                specs_map[name] = opts
    except Exception:
        specs_map = {}

    # 构建选项文本 -> 索引 的映射（加速查找）
    option_index_map = {}
    try:
        for dim in dim_names:
            options = specs_map.get(dim, [])
            option_index_map[dim] = {opt: idx for idx, opt in enumerate(options)}
    except Exception:
        option_index_map = {}

    # 简单的YAML转义（尽量减少依赖）
    def _yaml_escape_value(val: str) -> str:
        try:
            s = '' if val is None else str(val)
        except Exception:
            s = ''
        needs_quote = False
        if not s:
            needs_quote = True
        if (s.strip() != s) or any(ch in s for ch in [':', '#', '{', '}', '[', ']', ',', '&', '*', '!', '|', '>', "'", '"', '%', '@', '`']):
            needs_quote = True
        lowered = s.lower()
        if lowered in ("null", "true", "false", "yes", "no", "on", "off"):
            needs_quote = True
        if needs_quote:
            s = s.replace('\\', r'\\').replace('"', r'\"')
            return f'"{s}"'
        return s

    def _yaml_escape_key(key: str) -> str:
        # key 与 value 同步处理，尽量安全
        return _yaml_escape_value(key)

    # 提取价格：返回 (price_float_or_none, price_text)
    def _extract_price(val) -> tuple:
        try:
            if isinstance(val, (int, float)):
                return float(val), f"{val}"
        except Exception:
            pass
        try:
            s = str(val or '').strip()
        except Exception:
            s = ''
        # 同 Excel 逻辑：抽取首个数字片段
        try:
            tokens = ''.join((c if (c.isdigit() or c in '.,') else ' ') for c in s).split()
            for t in tokens:
                if any(ch.isdigit() for ch in t):
                    try:
                        return float(t.replace(',', '')), s
                    except Exception:
                        continue
        except Exception:
            pass
        return None, s

    # 去重“规格图”并生成友好文件名（与 Excel 下载/命名规则一致）
    out_dir = file_path.parent
    spec_img_urls, friendly_map = _generate_spec_image_filenames(headers, results, out_dir)

    # 下载规格图到本地（与友好文件名一致保存为 PNG；在缺少 Pillow 时以原始字节落盘）
    downloaded = 0
    if spec_img_urls:
        for u in spec_img_urls:
            try:
                fname = friendly_map.get(u) or f"图片_{hashlib.md5(u.encode('utf-8')).hexdigest()[:10]}.png"
                target = out_dir / fname
                if target.exists():
                    continue
                # 若 requests 不可用，跳过下载
                if requests is None:
                    continue
                req_headers = {
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0"
                    ),
                    "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
                    "Accept-Language": "zh-CN,zh;q=0.9",
                    "Referer": f"{urlparse(u).scheme}://{urlparse(u).hostname}/",
                    "Connection": "keep-alive",
                }
                resp = requests.get(u, headers=req_headers, timeout=20)
                resp.raise_for_status()
                if _HAS_PILLOW:
                    from io import BytesIO
                    bio = BytesIO(resp.content)
                    pil_img = PILImage.open(bio)
                    if pil_img.mode not in ("RGB", "RGBA"):
                        pil_img = pil_img.convert("RGB")
                    pil_img.save(target, format="PNG")
                else:
                    # 无 Pillow：直接按 PNG 后缀落盘原始字节（可能非 PNG，但保证文件存在）
                    with open(target, 'wb') as fbin:
                        fbin.write(resp.content)
                downloaded += 1
            except Exception:
                pass
        try:
            if downloaded:
                print(f"[步骤] 已下载规格图 {downloaded} 张到: {out_dir}")
        except Exception:
            pass

    # 生成 YAML 文本（紧凑形式 + 注释）
    lines: List[str] = []
    lines.append('# 本文件由 RPA 自动生成：存储淘宝商品规格与组合信息（紧凑表示）\n')
    lines.append('# 读取提示：先用 dims 与 specs 解析索引，再结合 combos 使用。\n')
    lines.append('specs:\n')
    for dim_name, options in specs_map.items():
        lines.append(f"  {_yaml_escape_key(dim_name)}:\n")
        for opt in options:
            lines.append(f"    - {_yaml_escape_value(opt)}\n")

    lines.append('\n# 维度名称顺序（与 specs 对齐，用于解释索引）\n')
    # 维度顺序
    lines.append('dims:\n')
    for dim in dim_names:
        lines.append(f"  - {_yaml_escape_key(dim)}\n")

    lines.append('\n# 商品主图（主图画廊，当前未专门采集，预留字段）\n')
    lines.append('product_images: []\n')
    lines.append('product_images_local: []\n')

    lines.append('\n# 规格图（主图区域在选中特定规格后展示的图片）\n')
    lines.append('spec_images:\n')
    for u in spec_img_urls:
        file_name = friendly_map.get(u) or f"图片_{hashlib.md5(u.encode('utf-8')).hexdigest()[:10]}.png"
        lines.append('  - file: ' + _yaml_escape_value(file_name) + '\n')
        lines.append('    url: ' + _yaml_escape_value(u) + '\n')

    lines.append('\n# 紧凑组合：每行 = [各维度选项索引..., 价格]\n')
    lines.append('# 说明：选项索引基于 specs[dims[i]]（0 起）；价格为数值或 null\n')
    # 紧凑组合：行内序列（flow style）：[d0, d1, ..., price]
    lines.append('combos:\n')
    for row in results or []:
        try:
            # 维度索引列表
            idx_list: List[str] = []
            for i in range(dim_count):
                dim_n = dim_names[i] if i < len(dim_names) else f"维度{i+1}"
                try:
                    v_txt = str(row[i]).strip()
                except Exception:
                    v_txt = ''
                idx = option_index_map.get(dim_n, {}).get(v_txt, -1)
                idx_list.append(str(idx))
            # 价格
            price_val = row[-1] if row else ''
            price_num, _price_text = _extract_price(price_val)
            price_repr = 'null' if price_num is None else (str(price_num))
            # 合成一行：-[idxs..., price]
            arr = ', '.join(idx_list + [price_repr])
            lines.append(f"  - [{arr}]\n")
        except Exception:
            continue

    # 写入文件
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    try:
        print(f"[步骤] 已导出YAML到: {file_path}")
    except Exception:
        pass
    try:
        append_to_log(f"导出YAML完成: {file_path}")
    except Exception:
        pass
    return file_path